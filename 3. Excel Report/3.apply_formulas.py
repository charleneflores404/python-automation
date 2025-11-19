from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
import os

# Get the directory where the script is located
script_dir = os.path.dirname(os.path.abspath(__file__))

# Paths for input and output files
input_path = os.path.join(script_dir,'barchart.xlsx')
output_path = os.path.join(script_dir, 'report.xlsx')

wb = load_workbook(input_path)
sheet = wb['Report']

min_column = wb.active.min_column
max_column = wb.active.max_column
min_row = wb.active.min_row
max_row = wb.active.max_row

# Write an Excel formula with Python
# sheet['B8'] = '=SUM(B6:B7)'
# sheet['B8'].style = 'Currency'

# Write multiple formulas with a for loop
for i in range(min_column+1, max_column+1):  # (B, G+1)
    col_letter = get_column_letter(i)
    sheet[f'{col_letter}{max_row + 1}'] = f'=SUM({col_letter}{min_row + 1}:{col_letter}{max_row})'
    sheet[f'{col_letter}{max_row + 1}'].style = 'Currency'
    sheet.column_dimensions[col_letter].width = 14  # make sure the values fit, otherwise you will see: #####

wb.save(output_path)